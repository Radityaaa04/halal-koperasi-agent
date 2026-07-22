#!/usr/bin/env python3
"""
Communication Agent for HALAL Koperasi Multi-Agent System
Generates PDF reports and Excel checklists for audit results.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    PageBreak,
)

from halal_koperasi_agent.config import settings
from halal_koperasi_agent.schemas.audit import (
    AuditCategory,
    AuditFinding,
    AuditSimulationResult,
    PriorityAction,
)
from halal_koperasi_agent.schemas.documents import DocumentType, DocumentMetadata
from halal_koperasi_agent.schemas.regulatory import RAGAnswer


class CommunicationAgent:
    """
    Agent responsible for:
    1. Generating comprehensive PDF audit report
    2. Generating Excel checklist with findings
    3. Creating executive summary for decision makers
    """

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles for PDF"""
        self.styles.add(ParagraphStyle(
            name='TitleIndo',
            parent=self.styles['Title'],
            fontSize=18,
            leading=22,
            alignment=TA_CENTER,
            spaceAfter=12,
            fontName='Helvetica-Bold',
        ))
        self.styles.add(ParagraphStyle(
            name='Heading1Indo',
            parent=self.styles['Heading1'],
            fontSize=14,
            leading=18,
            spaceBefore=12,
            spaceAfter=6,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#1B4F72'),
        ))
        self.styles.add(ParagraphStyle(
            name='Heading2Indo',
            parent=self.styles['Heading2'],
            fontSize=12,
            leading=16,
            spaceBefore=10,
            spaceAfter=4,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#2E86C1'),
        ))
        self.styles.add(ParagraphStyle(
            name='BodyIndo',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=4,
            fontName='Helvetica',
            alignment=TA_LEFT,
        ))
        self.styles.add(ParagraphStyle(
            name='BodyBold',
            parent=self.styles['Normal'],
            fontSize=10,
            leading=14,
            spaceAfter=4,
            fontName='Helvetica-Bold',
        ))
        self.styles.add(ParagraphStyle(
            name='TableCell',
            parent=self.styles['Normal'],
            fontSize=8,
            leading=10,
            fontName='Helvetica',
        ))
        self.styles.add(ParagraphStyle(
            name='TableCellBold',
            parent=self.styles['Normal'],
            fontSize=8,
            leading=10,
            fontName='Helvetica-Bold',
        ))

    def generate_pdf_report(
        self,
        result: AuditSimulationResult,
        documents: Dict[DocumentType, DocumentMetadata],
        rag_answers: List[RAGAnswer],
        output_path: Path,
    ) -> Path:
        """Generate comprehensive PDF audit report"""
        output_path.parent.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(
            str(output_path),
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        story = []

        # ========== COVER PAGE ==========
        story.extend(self._build_cover_page(result))
        story.append(PageBreak())

        # ========== EXECUTIVE SUMMARY ==========
        story.extend(self._build_executive_summary(result))
        story.append(PageBreak())

        # ========== DETAILED FINDINGS ==========
        story.extend(self._build_detailed_findings(result))
        story.append(PageBreak())

        # ========== CATEGORY SCORES ==========
        story.extend(self._build_category_scores(result))
        story.append(PageBreak())

        # ========== PRIORITY ACTIONS ==========
        story.extend(self._build_priority_actions(result))
        story.append(PageBreak())

        # ========== REGULATORY REFERENCES ==========
        story.extend(self._build_regulatory_refs(rag_answers))
        story.append(PageBreak())

        # ========== DOCUMENT INVENTORY ==========
        story.extend(self._build_document_inventory(documents))

        # Build PDF
        doc.build(story)
        logger.success(f"PDF report generated: {output_path}")
        return output_path

    def _build_cover_page(self, result: AuditSimulationResult) -> List:
        """Build cover page elements"""
        elements = []
        
        # Logo placeholder
        elements.append(Spacer(1, 4*cm))
        
        # Title
        elements.append(Paragraph(
            "LAPORAN SIMULASI AUDIT<br/>SERTIFIKASI HALAL",
            self.styles['TitleIndo']
        ))
        elements.append(Spacer(1, 1*cm))
        
        elements.append(Paragraph(
            f"<b>{result.koperasi_nama}</b>",
            ParagraphStyle('SubTitle', parent=self.styles['TitleIndo'], fontSize=16)
        ))
        elements.append(Spacer(1, 0.5*cm))
        
        # Info table
        info_data = [
            ["Application ID", result.application_id],
            ["Tanggal Simulasi", result.simulated_at.strftime("%d %B %Y")],
            ["Skor Kesiapan", f"{result.overall_readiness_score:.1f}%"],
            ["Rekomendasi", result.submission_recommendation],
            ["Versi Agent", result.agent_version],
        ]
        
        info_table = Table(info_data, colWidths=[5*cm, 10*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LINEBELOW', (0, 0), (-1, -2), 0.5, colors.grey),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 2*cm))
        
        # Recommendation badge
        rec_colors = {
            "READY": colors.green,
            "NEEDS_MINOR_FIXES": colors.orange,
            "NEEDS_MAJOR_WORK": colors.HexColor('#FF8C00'),
            "NOT_READY": colors.red,
        }
        badge_color = rec_colors.get(result.submission_recommendation, colors.grey)
        
        badge_table = Table(
            [[Paragraph(f"<b>{result.submission_recommendation}</b>", 
                        ParagraphStyle('Badge', parent=self.styles['Normal'], 
                                     fontSize=14, textColor=colors.white, alignment=TA_CENTER))]],
            colWidths=[8*cm]
        )
        badge_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), badge_color),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('ROUNDEDCORNERS', [5, 5, 5, 5]),
        ]))
        elements.append(badge_table)
        
        return elements

    def _build_executive_summary(self, result: AuditSimulationResult) -> List:
        """Build executive summary section"""
        elements = []
        elements.append(Paragraph("RINGKASAN EKSEKUTIF", self.styles['Heading1Indo']))
        elements.append(Spacer(1, 0.3*cm))
        
        # Summary text
        summary_text = f"""
        Simulasi audit halal untuk <b>{result.koperasi_nama}</b> (ID: {result.application_id}) 
        dilakukan pada {result.simulated_at.strftime("%d %B %Y")}. 
        Secara keseluruhan, koperasi mencapai skor kesiapan <b>{result.overall_readiness_score:.1f}%</b> 
        dari 81 item checklist yang dievaluasi.
        """
        elements.append(Paragraph(summary_text, self.styles['BodyIndo']))
        elements.append(Spacer(1, 0.3*cm))
        
        # Key metrics table
        metrics_data = [
            ["Metrik", "Nilai"],
            ["Total Checklist", str(result.total_checks)],
            ["Lulus (PASS)", str(result.passed)],
            ["Gagal (FAIL)", str(result.failed)],
            ["Peringatan (WARNING)", str(result.warnings)],
            ["Tidak Berlaku (N/A)", str(result.not_applicable)],
            ["Skor Kesiapan", f"{result.overall_readiness_score:.1f}%"],
            ["Rekomendasi", result.submission_recommendation],
            ["Estimasi Perbaikan", f"{result.estimated_time_to_fix_days} hari"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[8*cm, 6*cm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')]),
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Critical gaps summary
        if result.critical_gaps:
            elements.append(Paragraph("GAP KRITIS YANG PERLU DIPERBAIKI SEGERA:", self.styles['Heading2Indo']))
            for gap in result.critical_gaps[:10]:
                elements.append(Paragraph(f"• {gap}", self.styles['BodyIndo']))
            if len(result.critical_gaps) > 10:
                elements.append(Paragraph(f"... dan {len(result.critical_gaps) - 10} gap kritis lainnya", self.styles['BodyIndo']))
        
        return elements

    def _build_detailed_findings(self, result: AuditSimulationResult) -> List:
        """Build detailed findings by category"""
        elements = []
        elements.append(Paragraph("TEMUAN DETAIL PER KATEGORI", self.styles['Heading1Indo']))
        elements.append(Spacer(1, 0.3*cm))
        
        for cat in AuditCategory:
            cat_findings = [f for f in result.findings if f.category == cat]
            if not cat_findings:
                continue
            
            elements.append(Paragraph(cat.value, self.styles['Heading2Indo']))
            elements.append(Spacer(1, 0.2*cm))
            
            # Table header
            table_data = [["ID", "Item Checklist", "Status", "Bukti", "Gap/Severity"]]
            
            for f in cat_findings:
                status_color = {
                    "PASS": colors.green,
                    "FAIL": colors.red,
                    "WARNING": colors.orange,
                    "NOT_APPLICABLE": colors.grey,
                    "PENDING": colors.blue,
                }.get(f.status, colors.black)
                
                evidence_short = f.evidence_found[:80] + "..." if len(f.evidence_found) > 80 else f.evidence_found
                gap_short = f.gap_description[:80] + "..." if f.gap_description and len(f.gap_description) > 80 else (f.gap_description or "-")
                
                table_data.append([
                    Paragraph(f.checklist_item_id, self.styles['TableCell']),
                    Paragraph(f.checklist_item, self.styles['TableCell']),
                    Paragraph(f.status, ParagraphStyle('Status', parent=self.styles['TableCellBold'], textColor=status_color)),
                    Paragraph(evidence_short, self.styles['TableCell']),
                    Paragraph(gap_short, self.styles['TableCell']),
                ])
            
            col_widths = [1.2*cm, 5*cm, 1.5*cm, 4.5*cm, 4.8*cm]
            findings_table = Table(table_data, colWidths=col_widths, repeatRows=1)
            findings_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            ]))
            elements.append(findings_table)
            elements.append(Spacer(1, 0.4*cm))
        
        return elements

    def _build_category_scores(self, result: AuditSimulationResult) -> List:
        """Build category score summary"""
        elements = []
        elements.append(Paragraph("SKOR PER KATEGORI", self.styles['Heading1Indo']))
        elements.append(Spacer(1, 0.3*cm))
        
        table_data = [["Kategori", "Skor", "Status"]]
        for cat, score in result.category_scores.items():
            status = "✓ Baik" if score >= 70 else "⚠ Perlu Perbaikan" if score >= 40 else "✗ Kritis"
            table_data.append([cat, f"{score:.1f}%", status])
        
        cat_table = Table(table_data, colWidths=[10*cm, 3*cm, 4*cm])
        cat_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')]),
        ]))
        elements.append(cat_table)
        
        return elements

    def _build_priority_actions(self, result: AuditSimulationResult) -> List:
        """Build priority actions section"""
        elements = []
        elements.append(Paragraph("PLAN PERBAIKAN PRIORITAS (TOP 10)", self.styles['Heading1Indo']))
        elements.append(Spacer(1, 0.3*cm))
        
        if not result.priority_actions:
            elements.append(Paragraph("Tidak ada aksi prioritas yang diperlukan.", self.styles['BodyIndo']))
            return elements
        
        table_data = [["#", "Aksi", "Item ID", "Penanggung Jawab", "Deadline (hari)", "Estimasi Biaya"]]
        for pa in result.priority_actions:
            table_data.append([
                str(pa.priority),
                Paragraph(pa.action, self.styles['TableCell']),
                pa.checklist_item_id,
                pa.responsible,
                str(pa.deadline_days),
                pa.estimated_cost or "-",
            ])
        
        action_table = Table(table_data, colWidths=[0.8*cm, 7*cm, 2*cm, 2.5*cm, 2*cm, 3*cm])
        action_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (4, 0), (4, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF8E1')]),
        ]))
        elements.append(action_table)
        
        return elements

    def _build_regulatory_refs(self, rag_answers: List[RAGAnswer]) -> List:
        """Build regulatory references from RAG answers"""
        elements = []
        elements.append(Paragraph("REFERENSI REGULATORI (DARI RAG)", self.styles['Heading1Indo']))
        elements.append(Spacer(1, 0.3*cm))
        
        if not rag_answers:
            elements.append(Paragraph("Tidak ada referensi regulatori dari RAG.", self.styles['BodyIndo']))
            return elements
        
        table_data = [["Pertanyaan", "Jawaban Ringkas", "Confidence", "Sumber"]]
        for ans in rag_answers:
            sources = ", ".join(set(c.source.value for c in ans.citations)) if ans.citations else "-"
            table_data.append([
                Paragraph(ans.question[:60] + "...", self.styles['TableCell']),
                Paragraph(ans.answer[:80] + "...", self.styles['TableCell']),
                f"{ans.confidence:.0%}",
                Paragraph(sources, self.styles['TableCell']),
            ])
        
        rag_table = Table(table_data, colWidths=[4*cm, 6*cm, 1.5*cm, 5*cm], repeatRows=1)
        rag_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')]),
        ]))
        elements.append(rag_table)
        
        return elements

    def _build_document_inventory(self, documents: Dict[DocumentType, DocumentMetadata]) -> List:
        """Build document inventory section"""
        elements = []
        elements.append(Paragraph("INVENTARIS DOKUMEN YANG DIPROSES", self.styles['Heading1Indo']))
        elements.append(Spacer(1, 0.3*cm))
        
        table_data = [["Jenis Dokumen", "Status", "Kelengkapan", "File Size", "Issues"]]
        for doc_type, meta in documents.items():
            issues = len(meta.validation_issues)
            issue_text = f"{issues} issue(s)" if issues else "Tidak ada"
            table_data.append([
                doc_type.value,
                meta.status.value,
                f"{meta.completeness_score:.0%}",
                f"{meta.file_size_kb} KB",
                issue_text,
            ])
        
        doc_table = Table(table_data, colWidths=[4*cm, 2.5*cm, 2*cm, 2*cm, 6*cm])
        doc_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B4F72')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (3, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F8FF')]),
        ]))
        elements.append(doc_table)
        
        return elements

    def generate_excel_checklist(
        self,
        result: AuditSimulationResult,
        output_path: Path,
    ) -> Path:
        """Generate detailed Excel checklist with findings"""
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb = Workbook()
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1B4F72", end_color="1B4F72", fill_type="solid")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # ========== SHEET 1: SUMMARY ==========
        ws1 = wb.active
        ws1.title = "Ringkasan"
        
        summary_data = [
            ["LAPORAN SIMULASI AUDIT HALAL", ""],
            ["Koperasi", result.koperasi_nama],
            ["Application ID", result.application_id],
            ["Tanggal Simulasi", result.simulated_at.strftime("%d %B %Y")],
            ["Skor Kesiapan", f"{result.overall_readiness_score:.1f}%"],
            ["Rekomendasi", result.submission_recommendation],
            ["Estimasi Perbaikan", f"{result.estimated_time_to_fix_days} hari"],
            ["", ""],
            ["METRIK", "NILAI"],
            ["Total Checklist", result.total_checks],
            ["Lulus (PASS)", result.passed],
            ["Gagal (FAIL)", result.failed],
            ["Peringatan (WARNING)", result.warnings],
            ["Tidak Berlaku", result.not_applicable],
        ]
        
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment
                if row_idx == 1 or row_idx == 9:
                    cell.font = header_font
                    cell.fill = header_fill
                elif row_idx <= 8:
                    cell.font = Font(bold=True)
        
        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 40
        
        # ========== SHEET 2: CATEGORY SCORES ==========
        ws2 = wb.create_sheet("Skor Kategori")
        cat_headers = ["Kategori", "Skor (%)", "Status"]
        for col_idx, header in enumerate(cat_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment
        
        for row_idx, (cat, score) in enumerate(result.category_scores.items(), 2):
            status = "Baik" if score >= 70 else "Perlu Perbaikan" if score >= 40 else "Kritis"
            fill = pass_fill if score >= 70 else (warn_fill if score >= 40 else fail_fill)
            for col_idx, value in enumerate([cat, score, status], 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment
                cell.fill = fill
        
        ws2.column_dimensions['A'].width = 50
        ws2.column_dimensions['B'].width = 15
        ws2.column_dimensions['C'].width = 20
        
        # ========== SHEET 3: DETAILED FINDINGS ==========
        ws3 = wb.create_sheet("Temuan Detail")
        finding_headers = [
            "ID", "Kategori", "Sub Kategori", "Item Checklist",
            "Status", "Bukti Ditemukan", "Bukti Diharapkan",
            "Gap Description", "Severity", "Rekomendasi",
            "Auto Fixable", "Est Fix (Days)", "Penanggung Jawab",
            "Referensi Regulasi"
        ]
        
        for col_idx, header in enumerate(finding_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment
        
        for row_idx, f in enumerate(result.findings, 2):
            status_fills = {
                "PASS": pass_fill, "FAIL": fail_fill,
                "WARNING": warn_fill, "NOT_APPLICABLE": na_fill, "PENDING": warn_fill,
            }
            row_fill = status_fills.get(f.status, None)
            
            values = [
                f.checklist_item_id,
                f.category.value,
                f.checklist_item,  # sub_category not in finding, use item
                f.checklist_item,
                f.status,
                f.evidence_found,
                f.evidence_expected,
                f.gap_description or "",
                f.severity,
                f.recommended_action or "",
                "Ya" if f.auto_fixable else "Tidak",
                f.estimated_fix_days,
                f.responsible_party,
                f.requirement_source,
            ]
            
            for col_idx, value in enumerate(values, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment
                if row_fill:
                    cell.fill = row_fill
        
        # Set column widths
        ws3.column_dimensions['A'].width = 8
        ws3.column_dimensions['B'].width = 25
        ws3.column_dimensions['C'].width = 25
        ws3.column_dimensions['D'].width = 40
        ws3.column_dimensions['E'].width = 12
        ws3.column_dimensions['F'].width = 30
        ws3.column_dimensions['G'].width = 30
        ws3.column_dimensions['H'].width = 30
        ws3.column_dimensions['I'].width = 12
        ws3.column_dimensions['J'].width = 35
        ws3.column_dimensions['K'].width = 10
        ws3.column_dimensions['L'].width = 12
        ws3.column_dimensions['M'].width = 20
        ws3.column_dimensions['N'].width = 40
        
        # ========== SHEET 4: PRIORITY ACTIONS ==========
        ws4 = wb.create_sheet("Plan Perbaikan")
        action_headers = [
            "Prioritas", "Aksi", "Item ID", "Penanggung Jawab",
            "Deadline (Hari)", "Estimasi Biaya", "Dependencies"
        ]
        
        for col_idx, header in enumerate(action_headers, 1):
            cell = ws4.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment
        
        for row_idx, pa in enumerate(result.priority_actions, 2):
            values = [
                pa.priority,
                pa.action,
                pa.checklist_item_id,
                pa.responsible,
                pa.deadline_days,
                pa.estimated_cost or "",
                ", ".join(pa.dependencies) if pa.dependencies else "",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment
        
        ws4.column_dimensions['A'].width = 10
        ws4.column_dimensions['B'].width = 50
        ws4.column_dimensions['C'].width = 12
        ws4.column_dimensions['D'].width = 20
        ws4.column_dimensions['E'].width = 15
        ws4.column_dimensions['F'].width = 25
        ws4.column_dimensions['G'].width = 30
        
        # ========== SHEET 5: CRITICAL GAPS ==========
        ws5 = wb.create_sheet("Gap Kritis")
        gap_headers = ["No", "Gap Kritis", "Item Checklist ID", "Kategori", "Severity", "Est Fix (Days)"]
        for col_idx, header in enumerate(gap_headers, 1):
            cell = ws5.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = wrap_alignment
        
        for row_idx, gap in enumerate(result.critical_gaps, 2):
            # Find the finding for this gap
            finding = next((f for f in result.findings if f.checklist_item == gap), None)
            if finding:
                values = [row_idx-1, gap, finding.checklist_item_id, finding.category.value, finding.severity, finding.estimated_fix_days]
                for col_idx, value in enumerate(values, 1):
                    cell = ws5.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = wrap_alignment
                    cell.fill = fail_fill
        
        ws5.column_dimensions['A'].width = 5
        ws5.column_dimensions['B'].width = 50
        ws5.column_dimensions['C'].width = 15
        ws5.column_dimensions['D'].width = 25
        ws5.column_dimensions['E'].width = 12
        ws5.column_dimensions['F'].width = 15
        
        wb.save(output_path)
        logger.success(f"Excel checklist generated: {output_path}")
        return output_path

    def generate_all_reports(
        self,
        result: AuditSimulationResult,
        documents: Dict[DocumentType, DocumentMetadata],
        rag_answers: List[RAGAnswer],
        output_dir: Path,
    ) -> Dict[str, Path]:
        """Generate all report formats"""
        output_dir.mkdir(parents=True, exist_ok=True)
        
        pdf_path = output_dir / f"audit_report_{result.application_id}.pdf"
        excel_path = output_dir / f"audit_checklist_{result.application_id}.xlsx"
        
        self.generate_pdf_report(result, documents, rag_answers, pdf_path)
        self.generate_excel_checklist(result, excel_path)
        
        return {"pdf": pdf_path, "excel": excel_path}


# CLI for testing
if __name__ == "__main__":
    import sys
    from halal_koperasi_agent.schemas.documents import DocumentType, DocumentStatus
    from halal_koperasi_agent.schemas.audit import AuditCategory
    
    # Create mock result for testing
    mock_result = AuditSimulationResult(
        application_id="TEST-001",
        koperasi_nama="Koperasi Test",
        overall_readiness_score=65.0,
        total_checks=81,
        passed=30,
        failed=20,
        warnings=15,
        not_applicable=16,
        pending=0,
        findings=[
            AuditFinding(
                checklist_item_id="A01",
                checklist_item="Akta Pendirian Koperasi",
                category=AuditCategory.ADMINISTRASI,
                requirement_source="UU 25/2008",
                status="PASS",
                evidence_found="AKTA_PENDIRIAN.pdf (100%)",
                evidence_expected="Akta Pendirian",
                severity="INFO",
                auto_fixable=False,
                estimated_fix_days=0,
                responsible_party="Manajemen",
                regulatory_citations=[],
                agent_reasoning="Dokumen lengkap",
            ),
            AuditFinding(
                checklist_item_id="B01",
                checklist_item="Ketua HAS Terpenuhi",
                category=AuditCategory.HAS,
                requirement_source="BPJPH Peraturan 1/2023",
                status="FAIL",
                evidence_found="ORGANOGRAM_HAS.pdf (77%) - Sertifikat pelatihan missing",
                evidence_expected="Sertifikat pelatihan Ketua HAS",
                gap_description="Ketua HAS tidak memiliki sertifikat pelatihan valid",
                recommended_action="Daftarkan Ketua HAS ke pelatihan BPJPH/LPH",
                severity="CRITICAL",
                auto_fixable=False,
                estimated_fix_days=30,
                responsible_party="Ketua Koperasi",
                regulatory_citations=[],
                agent_reasoning="Mandatory requirement",
            ),
        ],
        critical_gaps=["Ketua HAS tidak memiliki sertifikat pelatihan valid"],
        high_priority_gaps=[],
        category_scores={
            "A. Persyaratan Administrasi": 85.0,
            "B. Halal Assurance System (HAS)": 20.0,
        },
        submission_recommendation="NEEDS_MAJOR_WORK",
        estimated_time_to_fix_days=60,
        priority_actions=[
            type('obj', (object,), {
                'priority': 1,
                'action': 'Daftarkan Ketua HAS ke pelatihan BPJPH/LPH',
                'checklist_item_id': 'B01',
                'responsible': 'Ketua Koperasi',
                'deadline_days': 30,
                'dependencies': [],
                'estimated_cost': 'Rp 5 Juta - Rp 15 Juta',
            })(),
        ],
    )
    
    # Test
    agent = CommunicationAgent()
    agent.generate_pdf_report(mock_result, {}, [], Path("output/test_report.pdf"))
    agent.generate_excel_checklist(mock_result, Path("output/test_checklist.xlsx"))
    print("Test reports generated!")